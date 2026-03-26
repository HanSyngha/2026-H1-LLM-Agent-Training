"""
종합 실습 솔루션: 검색 → 정리 → 저장 Agent

mini_agent_template.py의 완전한 솔루션입니다.
모든 TODO가 구현되어 있습니다.

=== 동작 흐름 ===
1. 사용자: "Python 3.12 새 기능을 검색하고 Excel로 저장해줘"
2. Agent → search_google("Python 3.12 새 기능")
3. Agent → get_page_content(검색결과_URL)
4. Agent → extract_and_format(페이지_텍스트, "Python 3.12")
5. Agent → save_to_excel("results.xlsx", JSON_데이터)
6. Agent → "검색 결과를 results.xlsx에 저장했습니다."
"""

import sys
import os
import json
import re
import requests

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))
from common.config import *

# ============================================================
# 라이브러리 임포트
# ============================================================

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print("경고: playwright가 없습니다. pip install playwright && playwright install chromium")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("경고: openpyxl이 없습니다. pip install openpyxl")


# ============================================================
# 1. Tool Registry
# ============================================================

class ToolRegistry:
    """도구 레지스트리 - 도구 등록, 스키마 조회, 실행"""

    def __init__(self):
        self._tools = {}

    def register(self, name: str, func, schema: dict):
        """도구를 등록합니다."""
        self._tools[name] = {"func": func, "schema": schema}

    def get_schemas(self) -> list[dict]:
        """모든 도구의 OpenAI Tool Schema를 반환합니다."""
        return [tool["schema"] for tool in self._tools.values()]

    def execute(self, name: str, arguments: dict) -> str:
        """도구를 실행합니다."""
        if name not in self._tools:
            return f"오류: 알 수 없는 도구 '{name}'"
        try:
            result = self._tools[name]["func"](**arguments)
            return str(result)
        except Exception as e:
            return f"도구 실행 오류 ({name}): {e}"


registry = ToolRegistry()


# ============================================================
# 2. 브라우저 매니저
# ============================================================

class BrowserManager:
    """브라우저 생명주기 관리 (싱글톤)"""

    _playwright = None
    _browser = None
    _page = None

    @classmethod
    def get_page(cls):
        """브라우저 페이지를 반환합니다. 필요 시 자동으로 시작합니다."""
        if not PLAYWRIGHT_AVAILABLE:
            raise RuntimeError("playwright가 설치되지 않았습니다.")

        if cls._page is None or cls._page.is_closed():
            if cls._playwright is None:
                cls._playwright = sync_playwright().start()
            if cls._browser is None or not cls._browser.is_connected():
                cls._browser = cls._playwright.chromium.launch(
                    headless=True,
                    args=["--no-sandbox", "--disable-dev-shm-usage"],
                )
            context = cls._browser.new_context(
                viewport={"width": 1280, "height": 720},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
            )
            cls._page = context.new_page()
        return cls._page

    @classmethod
    def close(cls):
        """브라우저를 종료합니다."""
        try:
            if cls._page and not cls._page.is_closed():
                cls._page.close()
            if cls._browser:
                cls._browser.close()
            if cls._playwright:
                cls._playwright.stop()
        except Exception:
            pass
        cls._page = None
        cls._browser = None
        cls._playwright = None


# ============================================================
# 3. 브라우저 도구 구현 (솔루션)
# ============================================================

def search_google(query: str) -> str:
    """
    구글에서 검색하고 상위 결과를 반환합니다.

    동작:
    1. 구글 검색 페이지로 이동
    2. JavaScript로 검색 결과(div.g) 추출
    3. 제목, URL, 설명을 정리하여 반환
    """
    try:
        page = BrowserManager.get_page()

        # 구글 검색 URL로 이동 (한국어 결과)
        search_url = f"https://www.google.com/search?q={query}&hl=ko"
        page.goto(search_url, wait_until="domcontentloaded", timeout=30000)

        # JavaScript로 검색 결과 추출
        results = page.evaluate("""
            () => {
                const items = document.querySelectorAll('div.g');
                return Array.from(items).slice(0, 10).map(item => {
                    const titleEl = item.querySelector('h3');
                    const linkEl = item.querySelector('a');
                    const descEl = item.querySelector('div[data-sncf], div.VwiC3b, span.aCOpRe');
                    return {
                        title: titleEl ? titleEl.innerText.trim() : '',
                        url: linkEl ? linkEl.href : '',
                        description: descEl ? descEl.innerText.trim() : ''
                    };
                }).filter(r => r.title && r.url);
            }
        """)

        if not results:
            # 대체 방법: 페이지 텍스트에서 추출 시도
            text = page.inner_text("body")
            return f"검색 결과를 구조화하여 추출하지 못했습니다.\n\n페이지 텍스트 (처음 3000자):\n{text[:3000]}"

        # 결과 문자열 정리
        output = f"=== 구글 검색 결과: '{query}' ({len(results)}건) ===\n\n"
        for i, r in enumerate(results, 1):
            output += f"{i}. {r['title']}\n"
            output += f"   URL: {r['url']}\n"
            if r['description']:
                output += f"   설명: {r['description'][:200]}\n"
            output += "\n"

        return output

    except Exception as e:
        return f"구글 검색 오류: {e}"


# 검색 도구 등록
registry.register("search_google", search_google, {
    "type": "function",
    "function": {
        "name": "search_google",
        "description": "구글에서 키워드를 검색하고 상위 결과(제목, URL, 설명)를 반환합니다.",
        "parameters": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "검색할 키워드",
                },
            },
            "required": ["query"],
        },
    },
})


def get_page_content(url: str) -> str:
    """
    웹 페이지의 텍스트 내용을 추출합니다.

    동작:
    1. 지정 URL로 이동
    2. body의 텍스트 추출
    3. 공백/줄바꿈 정리
    4. 길이 제한 적용
    """
    try:
        page = BrowserManager.get_page()

        # 페이지 이동
        page.goto(url, wait_until="domcontentloaded", timeout=30000)

        title = page.title()
        current_url = page.url

        # 텍스트 추출
        text = page.inner_text("body")

        # 텍스트 정리: 연속 공백/줄바꿈 축소
        text = re.sub(r"\n{3,}", "\n\n", text)
        text = re.sub(r" {2,}", " ", text)
        text = text.strip()

        # 길이 제한
        max_length = 8000
        truncated = False
        if len(text) > max_length:
            text = text[:max_length]
            truncated = True

        result = f"=== 페이지 내용 ===\n"
        result += f"URL: {current_url}\n"
        result += f"타이틀: {title}\n"
        result += f"{'=' * 40}\n\n"
        result += text

        if truncated:
            result += f"\n\n... (텍스트가 {max_length}자에서 잘렸습니다)"

        return result

    except Exception as e:
        return f"페이지 내용 추출 오류: {e}"


registry.register("get_page_content", get_page_content, {
    "type": "function",
    "function": {
        "name": "get_page_content",
        "description": "지정된 URL의 웹 페이지 텍스트 내용을 추출합니다.",
        "parameters": {
            "type": "object",
            "properties": {
                "url": {
                    "type": "string",
                    "description": "내용을 추출할 웹 페이지 URL",
                },
            },
            "required": ["url"],
        },
    },
})


# ============================================================
# 4. 데이터 추출 도구 (솔루션)
# ============================================================

def extract_and_format(raw_text: str, topic: str) -> str:
    """
    원시 텍스트에서 주제와 관련된 정보를 추출하고 정리합니다.

    간단한 구현: 키워드 기반 문장 필터링 + 정리
    (고급 구현: LLM sub-call로 더 정확한 추출 가능)
    """
    try:
        # 키워드 기반 관련 문장 추출
        keywords = topic.lower().split()
        lines = raw_text.split("\n")

        relevant_lines = []
        for line in lines:
            line_stripped = line.strip()
            if not line_stripped or len(line_stripped) < 10:
                continue
            # 키워드가 하나라도 포함된 줄을 선택
            line_lower = line_stripped.lower()
            if any(kw in line_lower for kw in keywords):
                relevant_lines.append(line_stripped)

        if not relevant_lines:
            # 키워드 매칭 실패 시 처음 50줄 반환
            relevant_lines = [
                l.strip() for l in lines
                if l.strip() and len(l.strip()) > 10
            ][:50]

        # 중복 제거 (순서 유지)
        seen = set()
        unique_lines = []
        for line in relevant_lines:
            if line not in seen:
                seen.add(line)
                unique_lines.append(line)

        result = f"=== '{topic}' 관련 정보 추출 ({len(unique_lines)}건) ===\n\n"
        for i, line in enumerate(unique_lines[:30], 1):  # 최대 30건
            result += f"{i}. {line}\n"

        if len(unique_lines) > 30:
            result += f"\n... 외 {len(unique_lines) - 30}건"

        return result

    except Exception as e:
        return f"데이터 추출 오류: {e}"


registry.register("extract_and_format", extract_and_format, {
    "type": "function",
    "function": {
        "name": "extract_and_format",
        "description": "원시 텍스트에서 주제와 관련된 정보를 추출하여 구조화합니다.",
        "parameters": {
            "type": "object",
            "properties": {
                "raw_text": {
                    "type": "string",
                    "description": "원시 텍스트 데이터",
                },
                "topic": {
                    "type": "string",
                    "description": "추출할 주제/키워드",
                },
            },
            "required": ["raw_text", "topic"],
        },
    },
})


# ============================================================
# 5. Excel 저장 도구 (솔루션)
# ============================================================

def save_to_excel(path: str, data: str) -> str:
    """
    데이터를 Excel 파일로 저장합니다.

    동작:
    1. JSON 문자열을 파싱
    2. openpyxl로 워크북 생성
    3. 헤더 + 데이터 입력
    4. 스타일 적용 (헤더 배경색, 볼드, 열 너비)
    5. 파일 저장
    """
    if not OPENPYXL_AVAILABLE:
        return "오류: openpyxl이 설치되지 않았습니다."

    try:
        # JSON 파싱
        parsed_data = json.loads(data)

        if not isinstance(parsed_data, list) or len(parsed_data) == 0:
            return "오류: 데이터는 비어있지 않은 배열이어야 합니다."

        abs_path = os.path.abspath(path)

        # 디렉토리 생성
        dir_path = os.path.dirname(abs_path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "검색 결과"

        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if isinstance(parsed_data[0], dict):
            # 딕셔너리 배열 형태
            headers = list(parsed_data[0].keys())

            # 헤더 행 추가
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 데이터 행 추가
            for row_data in parsed_data:
                row = [row_data.get(h, "") for h in headers]
                ws.append(row)

            # 데이터 셀에도 테두리 적용
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        elif isinstance(parsed_data[0], list):
            # 2차원 배열 형태
            for row_idx, row_data in enumerate(parsed_data):
                ws.append(row_data)
                if row_idx == 0:
                    for col_idx in range(1, len(row_data) + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

        else:
            return "오류: 지원하지 않는 데이터 형식입니다. 딕셔너리 배열 또는 2차원 배열을 사용하세요."

        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    cell_str = str(cell.value)
                    # 한글 보정 (더 넓은 폭)
                    korean_chars = sum(1 for c in cell_str if ord(c) > 127)
                    length = len(cell_str) + korean_chars
                    max_length = max(max_length, length)
            adjusted_width = min(max_length + 4, 60)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)

        # 첫 행 고정 (스크롤해도 헤더 보임)
        ws.freeze_panes = "A2"

        # 자동 필터 설정
        ws.auto_filter.ref = ws.dimensions

        # 저장
        wb.save(abs_path)

        file_size = os.path.getsize(abs_path)
        row_count = ws.max_row - 1  # 헤더 제외
        col_count = ws.max_column

        return (
            f"Excel 파일 저장 완료!\n"
            f"  경로: {abs_path}\n"
            f"  크기: {file_size:,} bytes\n"
            f"  데이터: {row_count}행 x {col_count}열\n"
            f"  시트: {ws.title}\n"
            f"  기능: 헤더 고정, 자동 필터, 스타일 적용"
        )

    except json.JSONDecodeError as e:
        return f"JSON 파싱 오류: {e}"
    except Exception as e:
        return f"Excel 저장 오류: {e}"


registry.register("save_to_excel", save_to_excel, {
    "type": "function",
    "function": {
        "name": "save_to_excel",
        "description": "데이터를 Excel 파일(.xlsx)로 저장합니다. 헤더 스타일, 자동 필터, 열 고정이 적용됩니다.",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "저장할 Excel 파일 경로 (예: 'results.xlsx')",
                },
                "data": {
                    "type": "string",
                    "description": "JSON 문자열. 딕셔너리 배열 형태. 예: '[{\"제목\":\"...\",\"URL\":\"...\",\"설명\":\"...\"}]'",
                },
            },
            "required": ["path", "data"],
        },
    },
})


# ============================================================
# 6. Agent Loop
# ============================================================

def call_llm(messages: list[dict], tools: list[dict] | None = None) -> dict:
    """OpenAI 호환 API 호출"""
    url = f"{GATEWAY_BASE_URL}/chat/completions"
    headers = get_headers()

    payload = {
        "model": DEFAULT_MODEL,
        "messages": messages,
    }
    if tools:
        payload["tools"] = tools

    response = requests.post(
        url,
        headers=headers,
        json=payload,
        proxies=PROXIES,
        timeout=120,
    )
    response.raise_for_status()
    return response.json()


SYSTEM_PROMPT = """당신은 웹 검색과 데이터 정리 전문 AI 어시스턴트입니다.

## 사용 가능한 도구
1. **search_google(query)**: 구글에서 키워드를 검색하고 상위 결과를 반환합니다.
2. **get_page_content(url)**: 웹 페이지의 텍스트 내용을 추출합니다.
3. **extract_and_format(raw_text, topic)**: 텍스트에서 주제 관련 정보를 추출합니다.
4. **save_to_excel(path, data)**: 데이터를 Excel 파일로 저장합니다.

## 작업 프로세스
사용자가 검색 및 저장을 요청하면 다음 순서로 작업하세요:

1. **검색**: search_google로 키워드 검색
2. **상세 확인**: 유용해 보이는 결과 1-3개의 URL을 get_page_content로 확인
3. **정보 정리**: 수집한 정보를 바탕으로 구조화된 데이터 생성
4. **Excel 저장**: save_to_excel로 데이터를 Excel 파일에 저장

## 데이터 저장 형식
save_to_excel의 data 파라미터는 JSON 딕셔너리 배열 문자열이어야 합니다.

좋은 예:
'[{"번호":1,"제목":"Python 3.12 Release","URL":"https://...","요약":"새로운 기능 소개"}]'

## 주의사항
- 검색 결과가 불충분하면 다른 키워드로 재검색하세요.
- 페이지 내용이 너무 길면 핵심만 추출하세요.
- Excel 저장 시 의미 있는 컬럼명을 사용하세요.
- 항상 한국어로 응답하세요.
"""


def run_agent(user_message: str) -> str:
    """Agent Loop를 실행합니다."""
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user_message},
    ]

    tool_schemas = registry.get_schemas()
    max_iterations = 20

    for iteration in range(1, max_iterations + 1):
        print(f"\n--- 반복 #{iteration} ---")

        response_data = call_llm(messages, tools=tool_schemas)
        assistant_message = response_data["choices"][0]["message"]
        tool_calls = assistant_message.get("tool_calls")

        if not tool_calls:
            # 최종 응답
            content = assistant_message.get("content", "")
            return content

        # 도구 호출 처리
        messages.append(assistant_message)

        for tc in tool_calls:
            tool_name = tc["function"]["name"]
            arguments = json.loads(tc["function"]["arguments"])
            tool_call_id = tc["id"]

            # 실행 정보 출력
            args_preview = json.dumps(arguments, ensure_ascii=False)
            if len(args_preview) > 120:
                args_preview = args_preview[:120] + "..."
            print(f"  [도구] {tool_name}({args_preview})")

            # 도구 실행
            result = registry.execute(tool_name, arguments)

            # 결과 미리보기
            preview = str(result)[:250].replace("\n", " ")
            if len(str(result)) > 250:
                preview += "..."
            print(f"  [결과] {preview}")

            messages.append({
                "role": "tool",
                "tool_call_id": tool_call_id,
                "content": str(result),
            })

    return "오류: 최대 반복 횟수를 초과했습니다."


# ============================================================
# 대화형 인터페이스
# ============================================================

def main():
    print("=" * 60)
    print("종합 실습 솔루션: 검색 → 정리 → 저장 Agent")
    print("=" * 60)
    print()
    print("사용 예시:")
    print("  - 'Python 3.12 새 기능을 검색하고 Excel로 저장해줘'")
    print("  - 'AI 트렌드 2024를 검색하고 결과를 정리해줘'")
    print("  - '서울 맛집을 검색하고 Excel에 정리해줘'")
    print()
    print("종료: 'quit' 또는 'exit'")
    print("=" * 60)

    try:
        while True:
            try:
                user_input = input("\n사용자: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\n\n종료합니다.")
                break

            if not user_input:
                continue

            if user_input.lower() in ("quit", "exit", "종료"):
                print("종료합니다.")
                break

            print("\nAgent:")
            result = run_agent(user_input)
            print(f"\n{result}")

    finally:
        # 브라우저 정리
        if PLAYWRIGHT_AVAILABLE:
            print("\n브라우저를 종료합니다...")
            BrowserManager.close()


if __name__ == "__main__":
    main()
