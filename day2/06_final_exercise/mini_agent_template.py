"""
종합 실습 템플릿: 검색 → 정리 → 저장 Agent

이 파일은 최종 실습의 시작 코드(starter code)입니다.
TODO 표시된 부분을 구현하여 완전한 Agent를 만드세요.

=== 목표 ===
1. 브라우저로 구글 검색
2. 검색 결과에서 정보 추출
3. 추출한 데이터를 Excel 파일로 저장

=== 아키텍처 ===
Agent Loop + Tool Registry + 브라우저 도구 + Excel 도구
"""

import sys
import os
import json
import requests

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))
from common.config import *

# ============================================================
# Playwright / openpyxl 임포트
# ============================================================

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print("경고: playwright가 없습니다. pip install playwright && playwright install chromium")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("경고: openpyxl이 없습니다. pip install openpyxl")


# ============================================================
# 1. Tool Registry (제공됨)
# ============================================================

class ToolRegistry:
    """간단한 도구 레지스트리"""

    def __init__(self):
        self._tools = {}  # {name: {"func": callable, "schema": dict}}

    def register(self, name: str, func, schema: dict):
        """도구를 등록합니다."""
        self._tools[name] = {"func": func, "schema": schema}

    def get_schemas(self) -> list[dict]:
        """모든 도구의 OpenAI Tool Schema를 반환합니다."""
        return [tool["schema"] for tool in self._tools.values()]

    def execute(self, name: str, arguments: dict) -> str:
        """도구를 실행합니다."""
        if name not in self._tools:
            return f"오류: 알 수 없는 도구 '{name}'"
        try:
            result = self._tools[name]["func"](**arguments)
            return str(result)
        except Exception as e:
            return f"도구 실행 오류: {e}"


# 레지스트리 인스턴스 생성
registry = ToolRegistry()


# ============================================================
# 2. 브라우저 매니저 (제공됨)
# ============================================================

class BrowserManager:
    """브라우저 생명주기 관리"""

    _playwright = None
    _browser = None
    _page = None

    @classmethod
    def get_page(cls):
        if cls._page is None or cls._page.is_closed():
            if cls._playwright is None:
                cls._playwright = sync_playwright().start()
            if cls._browser is None:
                cls._browser = cls._playwright.chromium.launch(headless=True)
            cls._page = cls._browser.new_page()
        return cls._page

    @classmethod
    def close(cls):
        try:
            if cls._page:
                cls._page.close()
            if cls._browser:
                cls._browser.close()
            if cls._playwright:
                cls._playwright.stop()
        except Exception:
            pass
        cls._page = None
        cls._browser = None
        cls._playwright = None


# ============================================================
# 3. 브라우저 도구 구현 (TODO)
# ============================================================

def search_google(query: str) -> str:
    """
    구글에서 검색하고 결과를 반환합니다.

    TODO: 다음을 구현하세요.
    1. BrowserManager.get_page()로 페이지 객체 획득
    2. 구글 검색 URL로 이동: https://www.google.com/search?q={query}&hl=ko
    3. 검색 결과 추출 (JavaScript evaluate 또는 selector 사용)
    4. 결과를 문자열로 정리하여 반환

    힌트:
    - page.goto(url, wait_until="domcontentloaded")
    - page.evaluate(javascript_code) 로 DOM에서 데이터 추출
    - 검색 결과는 div.g 셀렉터에 있습니다
    """
    # TODO: 여기에 구현하세요
    # ────────────────────────────────────────
    pass
    # ────────────────────────────────────────


# 검색 도구 스키마
SEARCH_SCHEMA = {
    "type": "function",
    "function": {
        "name": "search_google",
        "description": "구글에서 키워드를 검색하고 상위 결과를 반환합니다.",
        "parameters": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "검색할 키워드",
                },
            },
            "required": ["query"],
        },
    },
}

# 도구 등록
registry.register("search_google", search_google, SEARCH_SCHEMA)


def get_page_content(url: str) -> str:
    """
    웹 페이지의 텍스트 내용을 추출합니다.

    TODO: 다음을 구현하세요.
    1. BrowserManager.get_page()로 페이지 객체 획득
    2. page.goto(url)로 이동
    3. page.inner_text("body")로 텍스트 추출
    4. 텍스트를 적절히 정리 (연속 줄바꿈 제거 등)
    5. 최대 길이 제한 (예: 8000자)
    """
    # TODO: 여기에 구현하세요
    # ────────────────────────────────────────
    pass
    # ────────────────────────────────────────


PAGE_CONTENT_SCHEMA = {
    "type": "function",
    "function": {
        "name": "get_page_content",
        "description": "지정된 URL의 웹 페이지 텍스트 내용을 추출합니다.",
        "parameters": {
            "type": "object",
            "properties": {
                "url": {
                    "type": "string",
                    "description": "내용을 추출할 웹 페이지 URL",
                },
            },
            "required": ["url"],
        },
    },
}

registry.register("get_page_content", get_page_content, PAGE_CONTENT_SCHEMA)


# ============================================================
# 4. 데이터 추출 도구 (TODO)
# ============================================================

def extract_and_format(raw_text: str, topic: str) -> str:
    """
    원시 텍스트에서 주제와 관련된 정보를 추출하여 구조화합니다.

    TODO: 다음을 구현하세요.
    이 도구는 LLM이 추출한 정보를 정리하는 데 사용됩니다.
    간단한 구현으로는 텍스트를 그대로 반환해도 됩니다.

    고급 구현:
    - LLM을 한 번 더 호출하여 정보 추출 (sub-agent 패턴)
    - 정규표현식으로 패턴 매칭
    - 키워드 기반 필터링
    """
    # TODO: 여기에 구현하세요
    # ────────────────────────────────────────
    pass
    # ────────────────────────────────────────


EXTRACT_SCHEMA = {
    "type": "function",
    "function": {
        "name": "extract_and_format",
        "description": "원시 텍스트에서 주제와 관련된 정보를 추출하여 구조화합니다.",
        "parameters": {
            "type": "object",
            "properties": {
                "raw_text": {
                    "type": "string",
                    "description": "원시 텍스트 데이터",
                },
                "topic": {
                    "type": "string",
                    "description": "추출할 주제/키워드",
                },
            },
            "required": ["raw_text", "topic"],
        },
    },
}

registry.register("extract_and_format", extract_and_format, EXTRACT_SCHEMA)


# ============================================================
# 5. Excel 저장 도구 (TODO)
# ============================================================

def save_to_excel(path: str, data: str) -> str:
    """
    데이터를 Excel 파일로 저장합니다.

    Args:
        path: 저장할 파일 경로 (.xlsx)
        data: JSON 문자열 형태의 데이터
              예: '[{"제목":"...","URL":"...","설명":"..."}]'

    TODO: 다음을 구현하세요.
    1. data를 JSON으로 파싱
    2. openpyxl.Workbook()으로 워크북 생성
    3. 데이터를 시트에 입력 (헤더 포함)
    4. 헤더 스타일 적용 (볼드, 배경색)
    5. 열 너비 자동 조정
    6. 파일 저장

    힌트:
    - json.loads(data) 로 파싱
    - ws.append(row) 로 행 추가
    - Font(bold=True), PatternFill(...) 로 스타일
    """
    # TODO: 여기에 구현하세요
    # ────────────────────────────────────────
    pass
    # ────────────────────────────────────────


EXCEL_SCHEMA = {
    "type": "function",
    "function": {
        "name": "save_to_excel",
        "description": "데이터를 Excel 파일(.xlsx)로 저장합니다. 헤더 스타일이 자동 적용됩니다.",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "저장할 Excel 파일 경로 (예: 'results.xlsx')",
                },
                "data": {
                    "type": "string",
                    "description": "JSON 문자열. 딕셔너리 배열 형태. 예: '[{\"제목\":\"...\",\"URL\":\"...\"}]'",
                },
            },
            "required": ["path", "data"],
        },
    },
}

registry.register("save_to_excel", save_to_excel, EXCEL_SCHEMA)


# ============================================================
# 6. Agent Loop (제공됨)
# ============================================================

def call_llm(messages: list[dict], tools: list[dict] | None = None) -> dict:
    """OpenAI 호환 API 호출"""
    url = f"{GATEWAY_BASE_URL}/chat/completions"
    headers = get_headers()

    payload = {
        "model": DEFAULT_MODEL,
        "messages": messages,
    }
    if tools:
        payload["tools"] = tools

    response = requests.post(
        url,
        headers=headers,
        json=payload,
        proxies=PROXIES,
        timeout=120,
    )
    response.raise_for_status()
    return response.json()


def run_agent(user_message: str) -> str:
    """
    Agent Loop를 실행합니다. (제공됨)

    사용자 메시지를 받아:
    1. LLM에 전달
    2. tool_calls 확인 및 실행
    3. 최종 응답 반환
    """
    system_prompt = """당신은 웹 검색과 데이터 정리 전문 AI 어시스턴트입니다.

## 사용 가능한 도구
1. search_google(query): 구글 검색
2. get_page_content(url): 웹 페이지 텍스트 추출
3. extract_and_format(raw_text, topic): 텍스트에서 정보 추출
4. save_to_excel(path, data): 데이터를 Excel로 저장

## 작업 프로세스
1. 사용자의 검색 요청을 받으면 search_google로 검색
2. 유용한 결과의 상세 페이지를 get_page_content로 확인
3. 수집한 정보를 extract_and_format으로 정리
4. 정리된 데이터를 save_to_excel로 Excel 파일에 저장

## 데이터 저장 형식
save_to_excel에 전달하는 data는 JSON 딕셔너리 배열이어야 합니다.
예: '[{"제목":"Python 3.12","URL":"https://...","설명":"새로운 기능 소개"}]'

항상 한국어로 응답하세요.
"""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_message},
    ]

    tool_schemas = registry.get_schemas()
    max_iterations = 20

    for iteration in range(1, max_iterations + 1):
        print(f"\n--- 반복 #{iteration} ---")

        response_data = call_llm(messages, tools=tool_schemas)
        assistant_message = response_data["choices"][0]["message"]
        tool_calls = assistant_message.get("tool_calls")

        if not tool_calls:
            return assistant_message.get("content", "")

        messages.append(assistant_message)

        for tc in tool_calls:
            tool_name = tc["function"]["name"]
            arguments = json.loads(tc["function"]["arguments"])
            tool_call_id = tc["id"]

            print(f"  [도구] {tool_name}({json.dumps(arguments, ensure_ascii=False)[:100]})")

            result = registry.execute(tool_name, arguments)

            preview = str(result)[:200].replace("\n", " ")
            if len(str(result)) > 200:
                preview += "..."
            print(f"  [결과] {preview}")

            messages.append({
                "role": "tool",
                "tool_call_id": tool_call_id,
                "content": str(result),
            })

    return "오류: 최대 반복 횟수를 초과했습니다."


# ============================================================
# 메인 실행
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("종합 실습: 검색 → 정리 → 저장 Agent")
    print("=" * 60)
    print()
    print("TODO 부분을 구현한 후 실행하세요.")
    print()

    try:
        # 예시 실행
        query = "Python 3.12 새로운 기능을 검색하고 결과를 Excel로 저장해줘"
        print(f"[질문] {query}")
        print()

        result = run_agent(query)
        print(f"\n[최종 응답]\n{result}")

    finally:
        if PLAYWRIGHT_AVAILABLE:
            BrowserManager.close()
