"""
Office 자동화 도구 - Excel 파일 생성/읽기/수정

openpyxl 라이브러리를 사용하여 Excel 파일을 다루는 도구입니다.
LLM Agent가 데이터를 Excel로 정리하거나, 기존 Excel 파일을 분석할 수 있습니다.

=== 설치 ===
pip install openpyxl

=== Windows 앱 제어 (참고) ===
Windows 환경에서는 pywinauto, win32com 등을 통해 Office 앱을 직접 제어할 수 있습니다.
- pywinauto: GUI 자동화 (버튼 클릭, 메뉴 선택 등)
- win32com: COM 인터페이스 (Excel, Word, PowerPoint의 내부 API 호출)
- WSL에서는 Windows 앱을 직접 제어할 수 없으므로,
  파일 기반 접근(openpyxl)을 사용합니다.
"""

import sys
import os
import json
from typing import Any

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))
from common.config import *

try:
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("경고: openpyxl이 설치되지 않았습니다. 'pip install openpyxl'을 실행하세요.")


def _check_openpyxl():
    """openpyxl 설치 여부를 확인합니다."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl이 설치되지 않았습니다. 'pip install openpyxl'을 실행하세요.")


# ============================================================
# Excel 도구 함수
# ============================================================

def create_excel(path: str, data: str) -> str:
    """
    Excel 파일을 생성합니다.

    Args:
        path: 저장할 파일 경로 (.xlsx)
        data: JSON 문자열 형태의 데이터.
              형식 1 - 2차원 배열: '[["이름","점수"],["홍길동",95],["김철수",87]]'
              형식 2 - 딕셔너리 배열: '[{"이름":"홍길동","점수":95},{"이름":"김철수","점수":87}]'

    Returns:
        생성 결과 메시지
    """
    _check_openpyxl()

    try:
        abs_path = os.path.abspath(path)

        # 디렉토리 생성
        dir_path = os.path.dirname(abs_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)

        # JSON 데이터 파싱
        parsed_data = json.loads(data)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        if isinstance(parsed_data, list) and len(parsed_data) > 0:
            if isinstance(parsed_data[0], dict):
                # 딕셔너리 배열 형태: [{"이름": "홍길동", "점수": 95}, ...]
                headers = list(parsed_data[0].keys())
                ws.append(headers)

                # 헤더 스타일 적용
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # 데이터 행 추가
                for row_data in parsed_data:
                    ws.append([row_data.get(h, "") for h in headers])

            elif isinstance(parsed_data[0], list):
                # 2차원 배열 형태: [["이름", "점수"], ["홍길동", 95], ...]
                for row_idx, row_data in enumerate(parsed_data):
                    ws.append(row_data)

                    # 첫 번째 행을 헤더로 스타일 적용
                    if row_idx == 0:
                        for col_idx in range(1, len(row_data) + 1):
                            cell = ws.cell(row=1, column=col_idx)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")
            else:
                # 1차원 배열: 한 행으로 입력
                ws.append(parsed_data)

        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    # 한글은 폭이 넓으므로 보정
                    korean_chars = sum(1 for c in str(cell.value) if ord(c) > 127)
                    cell_length += korean_chars
                    max_length = max(max_length, cell_length)
            ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

        wb.save(abs_path)

        row_count = ws.max_row
        col_count = ws.max_column
        file_size = os.path.getsize(abs_path)

        return (
            f"Excel 파일 생성 완료\n"
            f"  경로: {abs_path}\n"
            f"  크기: {file_size:,} bytes\n"
            f"  행: {row_count}, 열: {col_count}"
        )

    except json.JSONDecodeError as e:
        return f"데이터 파싱 오류: JSON 형식이 올바르지 않습니다. {e}"
    except Exception as e:
        return f"Excel 생성 오류: {e}"


def read_excel(path: str) -> str:
    """
    Excel 파일을 읽어 내용을 반환합니다.

    Args:
        path: Excel 파일 경로 (.xlsx)

    Returns:
        파일 내용 (텍스트 테이블 형태)
    """
    _check_openpyxl()

    try:
        abs_path = os.path.abspath(path)

        if not os.path.exists(abs_path):
            return f"오류: 파일을 찾을 수 없습니다 - {abs_path}"

        wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)

        result = f"=== Excel 파일: {abs_path} ===\n"
        result += f"시트 목록: {wb.sheetnames}\n"

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result += f"\n--- 시트: {sheet_name} ---\n"
            result += f"행: {ws.max_row}, 열: {ws.max_column}\n\n"

            # 데이터 읽기 (최대 100행)
            rows_read = 0
            for row in ws.iter_rows(values_only=True):
                if rows_read >= 100:
                    result += f"\n... (100행까지만 표시, 전체 {ws.max_row}행)\n"
                    break
                # 각 셀을 문자열로 변환하여 탭으로 구분
                row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                result += row_str + "\n"
                rows_read += 1

        wb.close()
        return result

    except Exception as e:
        return f"Excel 읽기 오류: {e}"


def update_excel_cell(path: str, sheet: str, cell: str, value: str) -> str:
    """
    Excel 파일의 특정 셀을 업데이트합니다.

    Args:
        path: Excel 파일 경로
        sheet: 시트 이름
        cell: 셀 주소 (예: "A1", "B3")
        value: 새 값

    Returns:
        업데이트 결과
    """
    _check_openpyxl()

    try:
        abs_path = os.path.abspath(path)

        if not os.path.exists(abs_path):
            return f"오류: 파일을 찾을 수 없습니다 - {abs_path}"

        wb = openpyxl.load_workbook(abs_path)

        if sheet not in wb.sheetnames:
            return f"오류: 시트 '{sheet}'을 찾을 수 없습니다. 사용 가능한 시트: {wb.sheetnames}"

        ws = wb[sheet]
        old_value = ws[cell].value

        # 숫자로 변환 가능하면 숫자로 저장
        try:
            if "." in value:
                ws[cell] = float(value)
            else:
                ws[cell] = int(value)
        except (ValueError, TypeError):
            ws[cell] = value

        wb.save(abs_path)
        wb.close()

        return (
            f"셀 업데이트 완료\n"
            f"  파일: {abs_path}\n"
            f"  시트: {sheet}\n"
            f"  셀: {cell}\n"
            f"  이전 값: {old_value}\n"
            f"  새 값: {value}"
        )

    except Exception as e:
        return f"셀 업데이트 오류: {e}"


def create_excel_chart(path: str, data_range: str, chart_type: str = "bar") -> str:
    """
    Excel 파일에 차트를 추가합니다.

    Args:
        path: Excel 파일 경로
        data_range: 데이터 범위 (예: "A1:B10")
        chart_type: 차트 유형 - "bar"(막대), "line"(선), "pie"(원형)

    Returns:
        차트 생성 결과
    """
    _check_openpyxl()

    try:
        abs_path = os.path.abspath(path)

        if not os.path.exists(abs_path):
            return f"오류: 파일을 찾을 수 없습니다 - {abs_path}"

        wb = openpyxl.load_workbook(abs_path)
        ws = wb.active

        # 데이터 범위 파싱 (예: "A1:C10" → min_col=1, min_row=1, max_col=3, max_row=10)
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(data_range)

        # 차트 유형 선택
        chart_types = {
            "bar": BarChart,
            "line": LineChart,
            "pie": PieChart,
        }

        if chart_type not in chart_types:
            return f"오류: 지원하지 않는 차트 유형 '{chart_type}'. 사용 가능: {list(chart_types.keys())}"

        chart = chart_types[chart_type]()
        chart.title = "데이터 차트"
        chart.style = 10  # 깔끔한 스타일

        # 데이터 참조 설정
        # 첫 번째 열: 카테고리 (X축 레이블)
        # 나머지 열: 데이터 시리즈
        data = Reference(ws, min_col=min_col + 1, min_row=min_row,
                         max_col=max_col, max_row=max_row)
        categories = Reference(ws, min_col=min_col, min_row=min_row + 1,
                               max_row=max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # 차트 크기 설정
        chart.width = 15
        chart.height = 10

        # 차트를 시트에 추가 (데이터 오른쪽에 배치)
        chart_position = f"{get_column_letter(max_col + 2)}{min_row}"
        ws.add_chart(chart, chart_position)

        wb.save(abs_path)
        wb.close()

        return (
            f"차트 생성 완료\n"
            f"  파일: {abs_path}\n"
            f"  차트 유형: {chart_type}\n"
            f"  데이터 범위: {data_range}\n"
            f"  차트 위치: {chart_position}"
        )

    except Exception as e:
        return f"차트 생성 오류: {e}"


# ============================================================
# OpenAI Tool Schema 정의
# ============================================================

OFFICE_TOOL_SCHEMAS = [
    {
        "type": "function",
        "function": {
            "name": "create_excel",
            "description": "데이터를 Excel 파일(.xlsx)로 생성합니다. 헤더 스타일과 열 너비가 자동 적용됩니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "저장할 파일 경로 (예: 'output/report.xlsx')",
                    },
                    "data": {
                        "type": "string",
                        "description": (
                            "JSON 문자열 형태의 데이터. "
                            '형식1 - 2차원 배열: \'[["이름","점수"],["홍길동",95]]\'. '
                            '형식2 - 딕셔너리 배열: \'[{"이름":"홍길동","점수":95}]\''
                        ),
                    },
                },
                "required": ["path", "data"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_excel",
            "description": "Excel 파일의 내용을 읽어 텍스트로 반환합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "읽을 Excel 파일 경로 (.xlsx)",
                    },
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "update_excel_cell",
            "description": "Excel 파일의 특정 셀 값을 업데이트합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Excel 파일 경로",
                    },
                    "sheet": {
                        "type": "string",
                        "description": "시트 이름",
                    },
                    "cell": {
                        "type": "string",
                        "description": "셀 주소 (예: 'A1', 'B3')",
                    },
                    "value": {
                        "type": "string",
                        "description": "새 값",
                    },
                },
                "required": ["path", "sheet", "cell", "value"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_excel_chart",
            "description": "Excel 파일에 차트(막대/선/원형)를 추가합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Excel 파일 경로",
                    },
                    "data_range": {
                        "type": "string",
                        "description": "차트에 사용할 데이터 범위 (예: 'A1:C10')",
                    },
                    "chart_type": {
                        "type": "string",
                        "description": "차트 유형: 'bar'(막대), 'line'(선), 'pie'(원형). 기본값: 'bar'",
                        "enum": ["bar", "line", "pie"],
                    },
                },
                "required": ["path", "data_range"],
            },
        },
    },
]

# 도구 이름 -> 함수 매핑
OFFICE_TOOL_FUNCTIONS = {
    "create_excel": create_excel,
    "read_excel": read_excel,
    "update_excel_cell": update_excel_cell,
    "create_excel_chart": create_excel_chart,
}


# ============================================================
# 테스트
# ============================================================

if __name__ == "__main__":
    if not OPENPYXL_AVAILABLE:
        print("openpyxl이 설치되지 않아 테스트를 건너뜁니다.")
        sys.exit(1)

    print("=== Office 도구 테스트 ===\n")

    test_dir = "/tmp/office_test"
    os.makedirs(test_dir, exist_ok=True)

    # Excel 생성 (딕셔너리 배열)
    print("[create_excel - 딕셔너리 배열]")
    data = json.dumps([
        {"이름": "홍길동", "부서": "개발팀", "성과점수": 95},
        {"이름": "김철수", "부서": "마케팅팀", "성과점수": 87},
        {"이름": "이영희", "부서": "개발팀", "성과점수": 92},
        {"이름": "박민수", "부서": "디자인팀", "성과점수": 88},
    ], ensure_ascii=False)
    print(create_excel(f"{test_dir}/test1.xlsx", data))
    print()

    # Excel 생성 (2차원 배열)
    print("[create_excel - 2차원 배열]")
    data2 = json.dumps([
        ["월", "매출", "비용"],
        ["1월", 1000, 600],
        ["2월", 1200, 700],
        ["3월", 1500, 800],
    ], ensure_ascii=False)
    print(create_excel(f"{test_dir}/test2.xlsx", data2))
    print()

    # Excel 읽기
    print("[read_excel]")
    print(read_excel(f"{test_dir}/test2.xlsx"))
    print()

    # 셀 업데이트
    print("[update_excel_cell]")
    print(update_excel_cell(f"{test_dir}/test2.xlsx", "Sheet1", "B3", "1300"))
    print()

    # 차트 추가
    print("[create_excel_chart]")
    print(create_excel_chart(f"{test_dir}/test2.xlsx", "A1:C4", "bar"))
