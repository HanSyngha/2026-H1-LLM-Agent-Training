"""
Office 실습 정답: 데이터를 Excel로 생성하는 Agent

LLM Agent가 자연어 명령을 받아 데이터를 구조화하고,
openpyxl을 사용하여 Excel 파일을 생성합니다.

실행 방법:
    python exercise_solution.py

의존성:
    pip install openpyxl requests
"""

import json
import os
import sys
import requests

# 공통 설정 로드
import sys, os; sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))
from common.config import *

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("경고: openpyxl이 설치되지 않았습니다. pip install openpyxl")


# ============================================================
# 도구 구현: Excel 생성/읽기
# ============================================================

def create_excel(file_path: str, data: str) -> str:
    """데이터를 Excel 파일로 생성합니다.

    Args:
        file_path: 저장할 파일 경로 (.xlsx)
        data: JSON 문자열 형태의 데이터.
              형식 1 - 딕셔너리 배열: '[{"이름":"홍길동","점수":95}]'
              형식 2 - 2차원 배열: '[["이름","점수"],["홍길동",95]]'

    Returns:
        생성 결과 메시지
    """
    if not OPENPYXL_AVAILABLE:
        return "오류: openpyxl이 설치되지 않았습니다."

    try:
        parsed_data = json.loads(data)
    except json.JSONDecodeError as e:
        return f"오류: JSON 파싱 실패 - {e}"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "데이터"

        # 헤더 스타일을 정의합니다
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # 딕셔너리 배열 형식인지 확인합니다
        if isinstance(parsed_data, list) and parsed_data and isinstance(parsed_data[0], dict):
            headers = list(parsed_data[0].keys())
            # 헤더를 작성합니다
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 데이터를 작성합니다
            for row_idx, item in enumerate(parsed_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=item.get(header, ""))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

        # 2차원 배열 형식인 경우
        elif isinstance(parsed_data, list) and parsed_data and isinstance(parsed_data[0], list):
            for row_idx, row_data in enumerate(parsed_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    else:
                        cell.alignment = Alignment(horizontal="center")
        else:
            return "오류: 데이터 형식을 인식할 수 없습니다. 딕셔너리 배열 또는 2차원 배열을 사용하세요."

        # 열 너비를 자동 조정합니다
        for col in range(1, ws.max_column + 1):
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell_value = str(ws.cell(row=row, column=col).value or "")
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[get_column_letter(col)].width = max(max_length + 4, 10)

        # 파일을 저장합니다
        abs_path = os.path.abspath(file_path)
        os.makedirs(os.path.dirname(abs_path) if os.path.dirname(abs_path) else ".", exist_ok=True)
        wb.save(abs_path)

        return f"Excel 파일 생성 완료: {abs_path} (행: {ws.max_row}, 열: {ws.max_column})"

    except Exception as e:
        return f"Excel 생성 오류: {e}"


def read_excel(file_path: str) -> str:
    """Excel 파일을 읽어 내용을 반환합니다.

    Args:
        file_path: 읽을 Excel 파일 경로

    Returns:
        파일 내용 요약
    """
    if not OPENPYXL_AVAILABLE:
        return "오류: openpyxl이 설치되지 않았습니다."

    abs_path = os.path.abspath(file_path)
    if not os.path.exists(abs_path):
        return f"오류: 파일 없음 - {abs_path}"

    try:
        wb = openpyxl.load_workbook(abs_path, read_only=True)
        result = f"파일: {abs_path}\n시트 목록: {wb.sheetnames}\n\n"

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result += f"--- 시트: {sheet_name} (행: {ws.max_row}, 열: {ws.max_column}) ---\n"
            for row in ws.iter_rows(max_row=min(ws.max_row, 20), values_only=True):
                result += "  | ".join(str(cell or "") for cell in row) + "\n"
            if ws.max_row > 20:
                result += f"  ... (이하 {ws.max_row - 20}행 생략)\n"

        wb.close()
        return result

    except Exception as e:
        return f"Excel 읽기 오류: {e}"


# ============================================================
# Tool Schema 정의
# ============================================================

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "create_excel",
            "description": "데이터를 Excel 파일로 생성합니다. data는 JSON 문자열 (딕셔너리 배열 또는 2차원 배열)입니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": "저장할 파일 경로 (.xlsx)"},
                    "data": {"type": "string", "description": "JSON 문자열 형태의 데이터"},
                },
                "required": ["file_path", "data"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_excel",
            "description": "Excel 파일의 내용을 읽어 반환합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": "읽을 Excel 파일 경로"},
                },
                "required": ["file_path"],
            },
        },
    },
]

TOOL_FUNCTIONS = {
    "create_excel": create_excel,
    "read_excel": read_excel,
}


# ============================================================
# Agent Loop
# ============================================================

def call_llm(messages, tools=None):
    """LLM API를 호출합니다."""
    url = f"{GATEWAY_BASE_URL}/chat/completions"
    payload = {"model": DEFAULT_MODEL, "messages": messages}
    if tools:
        payload["tools"] = tools
    resp = requests.post(url, headers=get_headers(), json=payload,
                         proxies=PROXIES, timeout=120, verify=SSL_VERIFY)
    resp.raise_for_status()
    return resp.json()


def agent_loop(messages, max_iterations=10):
    """Agent Loop를 실행합니다."""
    for iteration in range(1, max_iterations + 1):
        data = call_llm(messages, tools=TOOLS)
        msg = data["choices"][0]["message"]
        tool_calls = msg.get("tool_calls")

        if not tool_calls:
            content = msg.get("content", "")
            messages.append({"role": "assistant", "content": content})
            return content

        messages.append(msg)
        for tc in tool_calls:
            name = tc["function"]["name"]
            args = json.loads(tc["function"]["arguments"])
            print(f"  [{iteration}] {name} 호출 중...")

            result = TOOL_FUNCTIONS.get(name, lambda **kw: "알 수 없는 도구")(**args)
            print(f"       -> {result[:150]}")

            messages.append({"role": "tool", "tool_call_id": tc["id"], "content": result})

    return "오류: 최대 반복 횟수 초과"


# ============================================================
# 대화형 인터페이스
# ============================================================

def main():
    """Office Agent 대화형 인터페이스를 실행합니다."""
    output_dir = os.path.dirname(os.path.abspath(__file__))

    print("=" * 60)
    print("  Office Agent (Excel 생성/읽기)")
    print("=" * 60)
    print(f"출력 경로: {output_dir}")
    print("예시: '직원 정보를 Excel로 만들어줘. 홍길동 개발팀, 김철수 마케팅팀'")
    print("'종료'로 끝내기\n")

    messages = [{
        "role": "system",
        "content": (
            f"당신은 데이터를 Excel 파일로 정리하는 AI 어시스턴트입니다. "
            f"사용자가 자연어로 설명한 데이터를 JSON으로 구조화하여 create_excel 도구로 Excel을 생성합니다. "
            f"파일은 '{output_dir}' 디렉토리에 저장하세요. "
            f"data 파라미터에는 반드시 유효한 JSON 문자열을 전달하세요. "
            f"한국어로 응답하세요."
        ),
    }]

    while True:
        try:
            user_input = input("\n사용자: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n종료합니다.")
            break

        if not user_input:
            continue
        if user_input.lower() in ["종료", "quit", "exit"]:
            break

        messages.append({"role": "user", "content": user_input})
        try:
            response = agent_loop(messages)
            print(f"\nAI: {response}")
        except Exception as e:
            print(f"\n[오류] {e}")


def demo_mode():
    """데모 모드: 미리 정의된 시나리오를 실행합니다."""
    output_dir = os.path.dirname(os.path.abspath(__file__))

    print("=" * 60)
    print("  Office Agent 데모")
    print("=" * 60)

    messages = [{
        "role": "system",
        "content": (
            f"당신은 데이터를 Excel 파일로 정리하는 AI 어시스턴트입니다. "
            f"사용자가 설명한 데이터를 JSON으로 구조화하여 create_excel 도구로 Excel을 생성하세요. "
            f"파일은 '{output_dir}' 디렉토리에 저장하세요. "
            f"data 파라미터에는 반드시 유효한 JSON 문자열을 전달하세요."
        ),
    }]

    # 데모 시나리오입니다
    query = (
        "다음 직원 정보를 Excel로 만들어줘. 파일명은 employees.xlsx로 해줘.\n"
        "홍길동, 개발팀, 연봉 5000만원\n"
        "김철수, 마케팅팀, 연봉 4500만원\n"
        "이영희, 디자인팀, 연봉 4800만원\n"
        "박민수, 데이터팀, 연봉 5200만원"
    )

    print(f"\n사용자: {query}")
    messages.append({"role": "user", "content": query})

    response = agent_loop(messages)
    print(f"\nAI: {response}")

    print(f"\n{'=' * 60}")
    print("  데모 완료!")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    if "--demo" in sys.argv:
        demo_mode()
    else:
        main()
