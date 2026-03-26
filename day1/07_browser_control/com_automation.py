"""
Windows COM 자동화 예제

COM (Component Object Model)은 Microsoft의 소프트웨어 컴포넌트 기술입니다.
Windows 애플리케이션 간의 프로세스 간 통신(IPC)을 가능하게 합니다.

=== COM 아키텍처 ===
- COM 서버: 기능을 제공하는 애플리케이션 (Excel, Word 등)
- COM 클라이언트: 기능을 사용하는 프로그램 (Python 스크립트)
- 인터페이스: IDispatch (late binding), 타입 라이브러리 (early binding)
- ProgID: 프로그램 식별자 (예: "Excel.Application", "Word.Application")
- CLSID: 고유 클래스 ID (레지스트리에 등록)

=== Python에서 COM 사용 ===
- win32com.client: pywin32 패키지의 COM 클라이언트 모듈
- Dispatch(): late binding (동적 호출)
- gencache.EnsureDispatch(): early binding (타입 정보 활용)
- DispatchEx(): 새 인스턴스 강제 생성

=== 주요 COM 서버 (ProgID) ===
- Excel.Application: Microsoft Excel
- Word.Application: Microsoft Word
- PowerPoint.Application: Microsoft PowerPoint
- Outlook.Application: Microsoft Outlook
- InternetExplorer.Application: Internet Explorer (레거시)
- Shell.Application: Windows Shell
- Scripting.FileSystemObject: 파일 시스템
- WScript.Shell: Windows Script Host

=== WSL 참고 ===
WSL에서는 COM을 직접 사용할 수 없습니다.
Windows에서 직접 실행하거나, WSL → Windows 호출 방법을 사용하세요.

=== 설치 ===
pip install pywin32 pywinauto  (Windows에서만)
"""

import sys
import os
import time
import platform

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))
from common.config import *


# ============================================================
# Windows 환경 감지 및 COM 라이브러리 임포트
# ============================================================

# WSL 또는 Linux에서 실행 시 graceful fallback
IS_WINDOWS = platform.system() == "Windows"
IS_WSL = "microsoft" in platform.uname().release.lower()

# COM 라이브러리 (Windows에서만 사용 가능)
try:
    import win32com.client
    import win32com.client.gencache
    import pythoncom
    COM_AVAILABLE = True
except ImportError:
    COM_AVAILABLE = False
    if IS_WINDOWS:
        print("경고: pywin32가 설치되지 않았습니다.")
        print("설치: pip install pywin32")
    else:
        print("참고: COM 자동화는 Windows에서만 사용 가능합니다.")
        print("현재 환경:", platform.system())
        if IS_WSL:
            print("WSL 환경이 감지되었습니다. Windows에서 직접 실행하세요.")

# pywinauto 라이브러리 (Windows GUI 자동화)
try:
    from pywinauto import Application as PywinautoApp
    from pywinauto import Desktop
    PYWINAUTO_AVAILABLE = True
except ImportError:
    PYWINAUTO_AVAILABLE = False
    if IS_WINDOWS:
        print("경고: pywinauto가 설치되지 않았습니다.")
        print("설치: pip install pywinauto")


# ============================================================
# Excel COM 자동화
# ============================================================

def excel_create_and_write():
    """
    COM을 사용하여 Excel을 제어합니다.

    win32com.client.Dispatch("Excel.Application")으로
    Excel COM 서버에 연결합니다.

    COM 호출 과정:
    1. Python → win32com → COM 런타임 → Excel.exe
    2. Excel.exe가 COM 서버로 동작하며 요청을 처리
    3. 결과가 역순으로 Python까지 전달
    """
    if not COM_AVAILABLE:
        print("[Excel] COM을 사용할 수 없습니다 (Windows에서 실행하세요)")
        return None

    excel = None
    try:
        print("[Excel] Excel 애플리케이션 시작...")

        # COM 스레드 초기화 (멀티스레드 환경에서 필요)
        pythoncom.CoInitialize()

        # Excel COM 서버 연결
        # Dispatch: late binding - 런타임에 메서드/속성 탐색
        excel = win32com.client.Dispatch("Excel.Application")

        # Excel 창 표시 여부 (True: 보임, False: 백그라운드)
        excel.Visible = True
        # 경고 다이얼로그 표시 여부
        excel.DisplayAlerts = False

        print("[Excel] 새 워크북 생성...")
        # Workbooks.Add(): 새 워크북(파일) 생성
        workbook = excel.Workbooks.Add()

        # ActiveSheet: 현재 활성 시트
        sheet = workbook.ActiveSheet
        sheet.Name = "자동화 테스트"

        print("[Excel] 데이터 입력...")

        # 셀에 값 입력 - Cells(행, 열) 또는 Range("A1")
        # 헤더 행
        headers = ["이름", "부서", "직급", "입사년도", "급여"]
        for col, header in enumerate(headers, 1):
            cell = sheet.Cells(1, col)
            cell.Value = header
            # 헤더 스타일 설정
            cell.Font.Bold = True
            cell.Font.Size = 12
            cell.Interior.Color = 0xFFCC99  # 배경색 (BGR 형식)

        # 데이터 행
        data = [
            ["김철수", "개발팀", "선임", 2019, 5500],
            ["이영희", "기획팀", "책임", 2017, 6500],
            ["박지민", "데이터팀", "수석", 2015, 7800],
            ["정수연", "AI팀", "선임", 2020, 5200],
            ["최동훈", "인프라팀", "책임", 2016, 7000],
        ]

        for row_idx, row_data in enumerate(data, 2):  # 2행부터 시작
            for col_idx, value in enumerate(row_data, 1):
                sheet.Cells(row_idx, col_idx).Value = value

        # 급여 열 서식 설정 (숫자 포맷)
        salary_range = sheet.Range(f"E2:E{len(data) + 1}")
        salary_range.NumberFormat = "#,##0"

        # 열 너비 자동 조정
        sheet.Columns.AutoFit()

        # 합계 행 추가
        sum_row = len(data) + 2
        sheet.Cells(sum_row, 4).Value = "합계:"
        sheet.Cells(sum_row, 4).Font.Bold = True
        # SUM 함수 사용
        sheet.Cells(sum_row, 5).Formula = f"=SUM(E2:E{len(data) + 1})"
        sheet.Cells(sum_row, 5).NumberFormat = "#,##0"
        sheet.Cells(sum_row, 5).Font.Bold = True

        # 평균 행 추가
        avg_row = sum_row + 1
        sheet.Cells(avg_row, 4).Value = "평균:"
        sheet.Cells(avg_row, 4).Font.Bold = True
        sheet.Cells(avg_row, 5).Formula = f"=AVERAGE(E2:E{len(data) + 1})"
        sheet.Cells(avg_row, 5).NumberFormat = "#,##0"

        print("[Excel] 차트 생성...")

        # 차트 추가
        chart_obj = sheet.ChartObjects().Add(
            Left=400,   # 가로 위치
            Top=50,     # 세로 위치
            Width=400,  # 너비
            Height=250, # 높이
        )
        chart = chart_obj.Chart
        # 차트 데이터 범위 설정
        chart.SetSourceData(sheet.Range(f"A1:E{len(data) + 1}"))
        chart.ChartType = 51  # xlBarClustered (가로 막대)
        chart.HasTitle = True
        chart.ChartTitle.Text = "부서별 급여"

        # 파일 저장
        save_path = os.path.join(os.path.expanduser("~"), "Desktop", "com_test.xlsx")
        print(f"[Excel] 파일 저장: {save_path}")
        workbook.SaveAs(save_path)

        print("[Excel] 완료!")
        print(f"  파일 경로: {save_path}")
        print(f"  시트 이름: {sheet.Name}")
        print(f"  데이터 행 수: {len(data)}")

        return save_path

    except Exception as e:
        print(f"[Excel] 오류: {e}")
        return None

    finally:
        # Excel 종료 (사용자가 확인할 수 있도록 잠시 대기)
        if excel:
            try:
                time.sleep(2)
                # 워크북 닫기 (저장 안 함 - 이미 저장됨)
                excel.Workbooks.Close()
                excel.Quit()
                print("[Excel] Excel 종료")
            except Exception:
                pass
        # COM 스레드 정리
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ============================================================
# Excel 데이터 읽기
# ============================================================

def excel_read_data(file_path: str):
    """
    COM을 사용하여 Excel 파일의 데이터를 읽습니다.

    COM 방식은 Excel이 설치되어 있어야 합니다.
    단순 데이터 읽기는 openpyxl이 더 가볍습니다.
    """
    if not COM_AVAILABLE:
        print("[Excel] COM을 사용할 수 없습니다")
        return []

    excel = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # 파일 열기
        abs_path = os.path.abspath(file_path)
        workbook = excel.Workbooks.Open(abs_path)
        sheet = workbook.ActiveSheet

        # 사용 범위 파악
        used_range = sheet.UsedRange
        rows = used_range.Rows.Count
        cols = used_range.Columns.Count

        print(f"[Excel] 데이터 범위: {rows}행 x {cols}열")

        # 데이터 읽기
        data = []
        for row in range(1, rows + 1):
            row_data = []
            for col in range(1, cols + 1):
                value = sheet.Cells(row, col).Value
                row_data.append(value)
            data.append(row_data)
            print(f"  행 {row}: {row_data}")

        workbook.Close(SaveChanges=False)
        return data

    except Exception as e:
        print(f"[Excel] 읽기 오류: {e}")
        return []

    finally:
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ============================================================
# Word COM 자동화
# ============================================================

def word_create_document():
    """
    COM을 사용하여 Word 문서를 생성합니다.

    Word.Application COM 서버를 통해
    문서 생성, 텍스트 입력, 서식 설정이 가능합니다.
    """
    if not COM_AVAILABLE:
        print("[Word] COM을 사용할 수 없습니다")
        return None

    word = None
    try:
        pythoncom.CoInitialize()
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = True
        word.DisplayAlerts = False

        print("[Word] 새 문서 생성...")
        doc = word.Documents.Add()

        # Selection 객체: 현재 커서 위치에 텍스트 입력
        selection = word.Selection

        # 제목 입력
        selection.Style = doc.Styles("제목 1")
        selection.TypeText("COM 자동화 보고서")
        selection.TypeParagraph()  # 줄바꿈

        # 본문 입력
        selection.Style = doc.Styles("본문")
        selection.TypeText("이 문서는 Python win32com을 사용하여 자동 생성되었습니다.")
        selection.TypeParagraph()
        selection.TypeParagraph()

        # 부제목
        selection.Style = doc.Styles("제목 2")
        selection.TypeText("1. COM 개요")
        selection.TypeParagraph()

        selection.Style = doc.Styles("본문")
        selection.TypeText(
            "COM(Component Object Model)은 Microsoft가 개발한 "
            "소프트웨어 컴포넌트 기술입니다. "
            "서로 다른 프로그래밍 언어로 작성된 소프트웨어가 "
            "상호 운용할 수 있게 합니다."
        )
        selection.TypeParagraph()

        # 파일 저장
        save_path = os.path.join(
            os.path.expanduser("~"), "Desktop", "com_test.docx"
        )
        doc.SaveAs(save_path)
        print(f"[Word] 문서 저장: {save_path}")

        return save_path

    except Exception as e:
        print(f"[Word] 오류: {e}")
        return None

    finally:
        if word:
            try:
                time.sleep(2)
                word.Documents.Close()
                word.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ============================================================
# pywinauto - Windows GUI 자동화
# ============================================================

def list_running_windows():
    """
    pywinauto를 사용하여 실행 중인 윈도우 목록을 조회합니다.

    pywinauto는 Windows UI 자동화 프레임워크입니다:
    - 윈도우 탐색 및 조작
    - 버튼 클릭, 텍스트 입력
    - 메뉴 접근
    - 대화상자 제어
    """
    if not PYWINAUTO_AVAILABLE:
        print("[pywinauto] 사용할 수 없습니다")
        return []

    try:
        print("[pywinauto] 실행 중인 윈도우 목록:")
        print("-" * 50)

        # Desktop 객체: 모든 최상위 윈도우에 접근
        desktop = Desktop(backend="uia")  # uia: UI Automation (최신)
        windows = desktop.windows()

        window_list = []
        for i, win in enumerate(windows, 1):
            try:
                title = win.window_text()
                if title:  # 제목이 있는 윈도우만
                    class_name = win.class_name()
                    rect = win.rectangle()
                    window_list.append({
                        "title": title,
                        "class": class_name,
                        "rect": f"({rect.left}, {rect.top}, {rect.right}, {rect.bottom})",
                    })
                    print(f"  {i}. {title}")
                    print(f"     클래스: {class_name}")
                    print(f"     위치: {rect}")
            except Exception:
                continue

        print(f"\n총 {len(window_list)}개의 윈도우")
        return window_list

    except Exception as e:
        print(f"[pywinauto] 오류: {e}")
        return []


def automate_notepad():
    """
    pywinauto를 사용하여 메모장을 자동으로 제어합니다.

    Application 클래스로 앱을 시작하거나 기존 앱에 연결합니다.
    backend 옵션:
    - "uia": UI Automation (권장, 최신 앱 지원)
    - "win32": Win32 API (레거시 앱 지원)
    """
    if not PYWINAUTO_AVAILABLE:
        print("[pywinauto] 사용할 수 없습니다")
        return

    try:
        print("[pywinauto] 메모장 자동화 시작...")

        # 메모장 실행
        app = PywinautoApp(backend="uia").start("notepad.exe")
        time.sleep(1)

        # 메모장 윈도우 찾기
        # 윈도우 제목 또는 클래스 이름으로 검색
        main_window = app.window(title_re=".*메모장.*")
        main_window.wait("ready", timeout=10)

        print("[pywinauto] 메모장 발견, 텍스트 입력...")

        # 편집 영역에 텍스트 입력
        edit = main_window.child_window(control_type="Edit")
        edit.type_keys(
            "COM 자동화 테스트{ENTER}"
            "pywinauto로 메모장을 제어하고 있습니다.{ENTER}"
            "이 텍스트는 자동으로 입력되었습니다.{ENTER}",
            with_spaces=True,
        )

        time.sleep(1)

        # 메뉴 접근: 파일 → 다른 이름으로 저장
        # menu_select로 메뉴 경로를 지정
        # main_window.menu_select("파일->다른 이름으로 저장")

        print("[pywinauto] 메모장 자동화 완료!")
        print("  (메모장은 열린 상태로 유지됩니다)")

    except Exception as e:
        print(f"[pywinauto] 오류: {e}")


# ============================================================
# 키 입력 전송 (SendKeys)
# ============================================================

def send_keystrokes_example():
    """
    키 입력을 특정 애플리케이션에 전송합니다.

    pywinauto의 type_keys() 특수 키:
    - {ENTER}: Enter 키
    - {TAB}: Tab 키
    - {ESC}: Escape 키
    - {BACKSPACE}: Backspace 키
    - {DELETE}: Delete 키
    - ^c: Ctrl+C (^는 Ctrl)
    - %f: Alt+F (%는 Alt)
    - +{TAB}: Shift+Tab (+는 Shift)
    - {F1}~{F12}: 펑션 키
    """
    if not PYWINAUTO_AVAILABLE:
        print("[SendKeys] pywinauto 사용 불가")
        return

    print("[SendKeys] 키 입력 전송 예제")
    print("  ^c = Ctrl+C")
    print("  ^v = Ctrl+V")
    print("  %{F4} = Alt+F4")
    print("  {ENTER} = Enter")
    print("  +{TAB} = Shift+Tab")

    # 실제로 키를 보내면 의도치 않은 동작이 발생할 수 있으므로
    # 코드 예시만 보여줍니다
    example_code = '''
    # 실행 중인 앱에 연결
    app = Application(backend="uia").connect(title="문서 제목")
    window = app.window(title="문서 제목")

    # 텍스트 입력
    window.type_keys("안녕하세요{ENTER}", with_spaces=True)

    # Ctrl+S (저장)
    window.type_keys("^s")

    # Ctrl+A (전체 선택) → Ctrl+C (복사)
    window.type_keys("^a")
    window.type_keys("^c")

    # Alt+F4 (종료)
    window.type_keys("%{F4}")
    '''
    print(f"\n코드 예시:\n{example_code}")


# ============================================================
# COM을 사용한 Windows Shell 제어
# ============================================================

def shell_automation_example():
    """
    Shell.Application COM 객체를 사용한 Windows Shell 제어 예제입니다.

    Shell COM으로 할 수 있는 작업:
    - 폴더 열기
    - 파일 탐색기 제어
    - 시스템 다이얼로그 표시
    - 프로그램 실행
    """
    if not COM_AVAILABLE:
        print("[Shell] COM을 사용할 수 없습니다")
        return

    try:
        pythoncom.CoInitialize()
        shell = win32com.client.Dispatch("Shell.Application")

        # 열린 탐색기 창 목록
        print("[Shell] 열린 탐색기 창 목록:")
        windows = shell.Windows()
        for i in range(windows.Count):
            try:
                win = windows.Item(i)
                if win:
                    location = win.LocationURL or win.LocationName
                    print(f"  {i+1}. {location}")
            except Exception:
                continue

        # 특수 폴더 경로 가져오기
        # 0: Desktop, 5: My Documents, 36: Windows, 37: System
        print("\n[Shell] 특수 폴더:")
        special_folders = {
            0: "바탕화면",
            5: "내 문서",
            36: "Windows",
            37: "System32",
        }
        namespace = shell.NameSpace
        for folder_id, name in special_folders.items():
            try:
                folder = namespace(folder_id)
                if folder:
                    print(f"  {name}: {folder.Self.Path}")
            except Exception:
                continue

    except Exception as e:
        print(f"[Shell] 오류: {e}")

    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ============================================================
# WScript.Shell - 프로세스 실행 및 환경 변수
# ============================================================

def wscript_shell_example():
    """
    WScript.Shell COM 객체를 사용한 시스템 제어 예제입니다.

    WScript.Shell로 할 수 있는 작업:
    - 프로그램 실행 (Run)
    - 환경 변수 읽기
    - 레지스트리 읽기/쓰기
    - 단축키 생성
    """
    if not COM_AVAILABLE:
        print("[WScript] COM을 사용할 수 없습니다")
        return

    try:
        pythoncom.CoInitialize()
        wshell = win32com.client.Dispatch("WScript.Shell")

        # 환경 변수 읽기
        env = wshell.Environment("Process")
        print("[WScript] 환경 변수 예시:")
        for var_name in ["USERNAME", "COMPUTERNAME", "OS", "PROCESSOR_ARCHITECTURE"]:
            try:
                value = env(var_name)
                print(f"  {var_name} = {value}")
            except Exception:
                print(f"  {var_name} = (접근 불가)")

        # 현재 디렉토리
        print(f"\n[WScript] 현재 디렉토리: {wshell.CurrentDirectory}")

        # 프로그램 실행 예시 (실제로 실행하지 않음)
        print("\n[WScript] 프로그램 실행 예시 (코드만):")
        print("  wshell.Run('notepad.exe')  # 메모장 실행")
        print("  wshell.Run('calc.exe')     # 계산기 실행")
        print("  wshell.Run('cmd /c dir')   # 명령 프롬프트")

    except Exception as e:
        print(f"[WScript] 오류: {e}")

    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ============================================================
# WSL 호환 대안 (COM 없이 사용 가능한 방법)
# ============================================================

def wsl_compatible_alternatives():
    """
    WSL/Linux 환경에서 COM 대신 사용할 수 있는 대안입니다.

    COM은 Windows 전용이므로, 크로스 플랫폼 대안을 알아둡니다.
    """
    print("=" * 60)
    print("WSL/Linux에서 사용 가능한 대안")
    print("=" * 60)

    alternatives = {
        "Excel 대안": {
            "openpyxl": "Excel 파일(.xlsx) 읽기/쓰기 (Excel 설치 불필요)",
            "pandas": "DataFrame을 Excel로 저장 (to_excel)",
            "xlsxwriter": "Excel 파일 생성 (서식, 차트 지원)",
        },
        "Word 대안": {
            "python-docx": "Word 문서(.docx) 생성 및 편집",
            "reportlab": "PDF 문서 생성",
        },
        "브라우저 자동화 대안": {
            "Playwright": "크로스 플랫폼 브라우저 자동화 (추천)",
            "Selenium": "크로스 플랫폼 브라우저 자동화 (레거시)",
            "CDP 직접": "Chrome DevTools Protocol (cdp_direct.py 참고)",
        },
        "GUI 자동화 대안": {
            "xdotool": "Linux X11 윈도우 제어",
            "xdg-open": "Linux 기본 앱으로 파일 열기",
            "subprocess": "Python 표준 라이브러리로 프로세스 제어",
        },
    }

    for category, tools in alternatives.items():
        print(f"\n  [{category}]")
        for tool_name, description in tools.items():
            print(f"    - {tool_name}: {description}")

    # openpyxl 예시 (COM 없이 Excel 생성)
    print("\n" + "-" * 40)
    print("openpyxl 예시 (COM 없이 Excel 파일 생성):")
    print("-" * 40)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "자동화 테스트"

        # 헤더
        headers = ["이름", "부서", "직급", "입사년도", "급여"]
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="FFCC99", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # 데이터
        data = [
            ["김철수", "개발팀", "선임", 2019, 5500],
            ["이영희", "기획팀", "책임", 2017, 6500],
            ["박지민", "데이터팀", "수석", 2015, 7800],
            ["정수연", "AI팀", "선임", 2020, 5200],
            ["최동훈", "인프라팀", "책임", 2016, 7000],
        ]

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 저장
        save_path = os.path.join(os.path.dirname(__file__), "openpyxl_test.xlsx")
        wb.save(save_path)
        print(f"  파일 저장 완료: {save_path}")

    except ImportError:
        print("  openpyxl이 설치되지 않았습니다.")
        print("  설치: pip install openpyxl")


# ============================================================
# 메인 실행
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("Windows COM 자동화 예제")
    print("=" * 60)
    print(f"플랫폼: {platform.system()}")
    print(f"WSL: {'예' if IS_WSL else '아니오'}")
    print(f"COM 사용 가능: {'예' if COM_AVAILABLE else '아니오'}")
    print(f"pywinauto 사용 가능: {'예' if PYWINAUTO_AVAILABLE else '아니오'}")
    print()

    if COM_AVAILABLE:
        # Windows에서 실행 시 COM 예제 실행
        print("[1] Excel COM 자동화")
        print("-" * 40)
        excel_create_and_write()
        print()

        print("[2] Shell COM 자동화")
        print("-" * 40)
        shell_automation_example()
        print()

        print("[3] WScript Shell")
        print("-" * 40)
        wscript_shell_example()
        print()

    if PYWINAUTO_AVAILABLE:
        print("[4] 실행 중인 윈도우 목록")
        print("-" * 40)
        list_running_windows()
        print()

        print("[5] 키 입력 전송 예제")
        print("-" * 40)
        send_keystrokes_example()
        print()

    # WSL/Linux 환경이거나 COM이 없는 경우 대안 표시
    if not COM_AVAILABLE or IS_WSL:
        print("[대안] WSL/Linux 호환 방법")
        print("-" * 40)
        wsl_compatible_alternatives()
